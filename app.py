import streamlit as st
from datetime import datetime, timedelta
import pandas as pd
import os
from openpyxl import load_workbook
import calendar
from dateutil.relativedelta import relativedelta
import subprocess
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# File paths
DATA_FILE = "tracker_data.xlsx"
HISTORY_FILE = "progress_history.xlsx"
PASSWORD = "param123"

# GitHub repository settings
REPO_DIR = "."  # Repository root (adjust if needed)
BRANCH = "main"  # Adjust to your branch name

# Load or initialize data
def load_data():
    if os.path.exists(DATA_FILE):
        try:
            df = pd.read_excel(DATA_FILE, engine='openpyxl')
            df['DateAdded'] = pd.to_datetime(df['DateAdded']).dt.date
            logger.info(f"Loaded {DATA_FILE} successfully")
        except Exception as e:
            logger.error(f"Error reading {DATA_FILE}: {e}")
            st.error(f"Error reading {DATA_FILE}: {e}")
            df = pd.DataFrame(columns=["GoalID", "GoalName", "Frequency", "Progress", "DateAdded", "Week"])
    else:
        logger.warning(f"{DATA_FILE} not found, initializing empty DataFrame")
        df = pd.DataFrame(columns=["GoalID", "GoalName", "Frequency", "Progress", "DateAdded", "Week"])
    return df

def load_history():
    if os.path.exists(HISTORY_FILE):
        try:
            df = pd.read_excel(HISTORY_FILE, engine='openpyxl')
            df['Date'] = pd.to_datetime(df['Date']).dt.date
            logger.info(f"Loaded {HISTORY_FILE} successfully")
        except Exception as e:
            logger.error(f"Error reading {HISTORY_FILE}: {e}")
            st.error(f"Error reading {HISTORY_FILE}: {e}")
            df = pd.DataFrame(columns=["GoalID", "GoalName", "Date", "Progress", "Percentage", "Change"])
    else:
        logger.warning(f"{HISTORY_FILE} not found, initializing empty DataFrame")
        df = pd.DataFrame(columns=["GoalID", "GoalName", "Date", "Progress", "Percentage", "Change"])
    return df

def save_data():
    try:
        # Save data to temporary Excel files without password
        temp_data_file = "temp_tracker_data.xlsx"
        temp_history_file = "temp_progress_history.xlsx"

        st.session_state.data.to_excel(temp_data_file, index=False, engine='openpyxl')
        st.session_state.history.to_excel(temp_history_file, index=False, engine='openpyxl')
        logger.info(f"Saved temporary files: {temp_data_file}, {temp_history_file}")

        # Load workbooks with openpyxl to set password
        data_wb = load_workbook(temp_data_file)
        history_wb = load_workbook(temp_history_file)

        # Set password protection
        data_wb.security.set_workbook_password(PASSWORD)
        history_wb.security.set_workbook_password(PASSWORD)

        # Save the password-protected files
        data_wb.save(DATA_FILE)
        history_wb.save(HISTORY_FILE)
        logger.info(f"Saved password-protected files: {DATA_FILE}, {HISTORY_FILE}")

        # Remove temporary files
        if os.path.exists(temp_data_file):
            os.remove(temp_data_file)
            logger.info(f"Removed temporary file: {temp_data_file}")
        if os.path.exists(temp_history_file):
            os.remove(temp_history_file)
            logger.info(f"Removed temporary file: {temp_history_file}")

        # Commit and push to GitHub
        commit_and_push_to_github()

    except Exception as e:
        logger.error(f"Error saving data: {e}")
        st.error(f"Error saving data: {e}")

def commit_and_push_to_github():
    try:
        # Ensure we're in the repository directory
        os.chdir(REPO_DIR)
        
        # Add files to git
        subprocess.run(["git", "add", DATA_FILE, HISTORY_FILE], check=True)
        logger.info(f"Added {DATA_FILE} and {HISTORY_FILE} to git")

        # Commit changes
        commit_message = f"Update tracker data {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        subprocess.run(["git", "commit", "-m", commit_message], check=True)
        logger.info(f"Committed changes: {commit_message}")

        # Push to GitHub
        subprocess.run(["git", "push", "origin", BRANCH], check=True)
        logger.info(f"Pushed changes to GitHub branch {BRANCH}")

    except subprocess.CalledProcessError as e:
        logger.error(f"Error committing/pushing to GitHub: {e}")
        st.error(f"Error committing to GitHub: {e}")
    except Exception as e:
        logger.error(f"Unexpected error during Git operations: {e}")
        st.error(f"Unexpected error during Git operations: {e}")

def get_week_number(date):
    return date.isocalendar()[1]

def add_goal(name, freq):
    goal_id = f"G{len(st.session_state.data) + 1}"
    today = datetime.now().date()
    week = get_week_number(today)
    new_row = {
        "GoalID": goal_id,
        "GoalName": name,
        "Frequency": freq,
        "Progress": 1.0,
        "DateAdded": today,
        "Week": week
    }
    hist_row = {
        "GoalID": goal_id,
        "GoalName": name,
        "Date": today,
        "Progress": 1.0,
        "Percentage": 0,
        "Change": 0
    }
    st.session_state.data = pd.concat([st.session_state.data, pd.DataFrame([new_row])], ignore_index=True)
    st.session_state.history = pd.concat([st.session_state.history, pd.DataFrame([hist_row])], ignore_index=True)
    save_data()
    st.success(f"Added Goal: {name}")
    st.rerun()

def update_progress(goal_id, goal_name, pct):
    today = datetime.now().date()
    current_week = get_week_number(today)

    # Get the goal's frequency
    frequency = st.session_state.data.loc[st.session_state.data["GoalID"] == goal_id, "Frequency"].values[0]

    # Check if an update is allowed
    if frequency == "Weekly":
        history_weeks = st.session_state.history[
            (st.session_state.history["GoalID"] == goal_id) & 
            (st.session_state.history["Date"].apply(get_week_number) == current_week)
        ]
        if not history_weeks.empty:
            st.error(f"Progress for this weekly goal has already been updated for week {current_week}.")
            return
    else:
        if not st.session_state.history[
            (st.session_state.history["GoalID"] == goal_id) & 
            (st.session_state.history["Date"] == today)
        ].empty:
            st.error("Progress for this goal has already been updated today.")
            return

    # Update progress
    current_progress = st.session_state.data.loc[st.session_state.data["GoalID"] == goal_id, "Progress"].values[0]
    if pct == 100:
        new_progress = current_progress * 1.01
        change = 0.01
    elif pct == 50:
        new_progress = current_progress * 1.005
        change = 0.005
    else:
        new_progress = current_progress / 1.01
        change = -0.01

    st.session_state.data.loc[st.session_state.data["GoalID"] == goal_id, "Progress"] = new_progress
    row = {
        "GoalID": goal_id,
        "GoalName": goal_name,
        "Date": today,
        "Progress": new_progress,
        "Percentage": pct,
        "Change": change
    }
    st.session_state.history = pd.concat([st.session_state.history, pd.DataFrame([row])], ignore_index=True)
    save_data()
    st.success("Progress updated!")
    st.rerun()

def delete_goal(goal_id, goal_name):
    st.session_state.data = st.session_state.data[st.session_state.data["GoalID"] != goal_id]
    st.session_state.history = st.session_state.history[st.session_state.history["GoalID"] != goal_id]
    save_data()
    st.success(f"Goal '{goal_name}' deleted successfully!")
    st.rerun()

def calculate_potential_progress(start_date, current_progress):
    today = datetime.now().date()
    days_since_start = (today - start_date).days
    potential_progress = 1.0
    for _ in range(days_since_start + 1):
        potential_progress *= 1.01
    return potential_progress

def show_progress_info(gid, gname):
    df = st.session_state.history[st.session_state.history["GoalID"] == gid]
    if df.empty:
        return
    progress = df["Progress"].iloc[-1] if not df.empty else 1.0
    start_date = st.session_state.data[st.session_state.data["GoalID"] == gid]["DateAdded"].iloc[0]
    potential_progress = calculate_potential_progress(start_date, progress)
    today = datetime.now().date()

    st.markdown(f"""
        <div style='
            display: flex;
            flex-direction: column;
            gap: 15px;
            margin-top: 10px;
        '>
            <div style='
                background: linear-gradient(45deg, #ff6b6b, #ff8e53);
                color: white;
                font-size: 18px;
                padding: 15px 20px;
                border-radius: 12px;
                font-weight: bold;
                box-shadow: 0 6px 12px rgba(0,0,0,0.3);
                text-align: center;
                transition: transform 0.2s ease-in-out;
            ' onmouseover='this.style.transform="scale(1.05)"' onmouseout='this.style.transform="scale(1)"'>
                🎯 Potential Progress: {potential_progress:.3f}
            </div>
            <div style='
                background: linear-gradient(45deg, #4facfe, #00f2fe);
                color: white;
                font-size: 18px;
                padding: 15px 20px;
                border-radius: 12px;
                font-weight: bold;
                box-shadow: 0 6px 12px rgba(0,0,0,0.3);
                text-align: center;
                transition: transform 0.2s ease-in-out;
            ' onmouseover='this.style.transform="scale(1.05)"' onmouseout='this.style.transform="scale(1)"'>
                🔥 Actual Progress: {progress:.3f}
            </div>
            <div style='
                background: linear-gradient(45deg, #2